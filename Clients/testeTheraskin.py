from openpyxl import load_workbook
from datetime import date
import cv2
import os

#pasta_excel = 'C:\\AutomationEdge\\Workspace\\output\\Imagens'
pasta_excel = df['pasta_download'].values[0]
pasta_download = os.path.join(pasta_excel,date.today().strftime('%d%m%Y'))  #Diretorio inicial {Entry.Current.Directory}
pasta_base = df['pasta_base'].values[0]
#pasta_base = r'C:\Base de Imagens'
data_arquivo = date.today().strftime('%Y%m%d')
nome_arquivo = f'ImagesReport_Output_{data_arquivo}.xlsx'
pasta_sem_coleta = rf'C:\AutomationEdge\Workspace\output\Imagens\{date.today().strftime("%d%m%Y")}'

def coleta_arquivos_baixados():
    arquivos = {}
    for pasta in os.listdir(pasta_download): #Coleta o nome da primeira pasta - Diretorio do Site
        if os.path.isdir(os.path.join(pasta_download, pasta)): #Verifica se é um diretório ou um arquivo
            arquivos[pasta] = [] #Cria lista para armazenar os produtos desse site
            for produtos in os.listdir(os.path.join(pasta_download, pasta)): #Coleta as subpastas para interagir
                infos = {} #Cria um dicionario temporario para adicionar a chave do site
                qtde_arquivos = os.listdir(os.path.join(pasta_download, pasta, produtos))
                if qtde_arquivos:
                    infos[produtos] = os.path.join(pasta_download, pasta, produtos, qtde_arquivos[0]) #Alimenta dicionario presente na variavel "infos"
                    arquivos[pasta].append(infos) #Adiciona o dicionario "info" no dicionario "arquivos"
    return arquivos

def coleta_nomes_imagens():
    base_imagens = os.listdir(pasta_base)
    return base_imagens

#Works well with images of different dimensions
def orb_sim(img1, img2):
    # SIFT is no longer available in cv2 so using ORB
    orb = cv2.ORB_create()

    # detect keypoints and descriptors
    kp_a, desc_a = orb.detectAndCompute(img1, None)
    kp_b, desc_b = orb.detectAndCompute(img2, None)

    # define the bruteforce matcher object
    bf = cv2.BFMatcher(cv2.NORM_HAMMING, crossCheck=True)
        
    #perform matches. 
    matches = bf.match(desc_a, desc_b)
    #Look for similar regions with distance < 50. Goes from 0 to 100 so pick a number between.
    similar_regions = [i for i in matches if i.distance < 50]  
    if len(matches) == 0:
        return 0
    return len(similar_regions) / len(matches)

def compara_imagens(arquivos):
    tabela = {'Empresa': [], 'Produto': [], 'Porcentagem': []}
    for nome_empresa in arquivos:
        for produto in arquivos[nome_empresa]:
            nome_produto = list(produto.keys())[0]
            for nome in arquivos_base:
                if nome_produto in nome:
                    img1 = cv2.imread(os.path.join(pasta_base, nome), 0)
                    img2 = cv2.imread(produto[nome_produto],0)
                    orb_similarity = orb_sim(img1, img2)  #1.0 smeans identical. Lower = not similar
                    #orb_similarity = float("{:.2f}".format(orb_similarity))
                    orb_similarity = orb_similarity * 100
                    orb_similarity = "{:.2f}".format(orb_similarity)
                    linhas = coleta_celula(nome_produto, nome_empresa)
                    for celula in linhas:
                        escreve_celula(celula, orb_similarity+'%')
                        #print(f'{nome_empresa}: {nome_produto}: {orb_similarity}')
                        tabela['Empresa'].append(nome_empresa)
                        tabela['Produto'].append(nome_produto)
                        tabela['Porcentagem'].append(orb_similarity)
    return tabela
    #return orb_similarity

def coleta_celula(produto_dic, empresa_dic):
    linhas = []
    wb = load_workbook(pasta_excel+'\\'+nome_arquivo)
    sheet = wb.active
    for row in range(1, sheet.max_row+1):
        produto = sheet.cell(row, 1).value
        empresa = sheet.cell(row, 2).value
        if produto_dic in produto and empresa_dic in empresa:
            linhas.append(row)
    wb.close()
    return linhas

def escreve_celula(linha, valor):
    wb = load_workbook(pasta_excel+'\\'+nome_arquivo)
    sheet = wb.active
    sheet.cell(linha, 6, valor)
    wb.save(pasta_excel+'\\'+nome_arquivo)

arquivos_baixados = coleta_arquivos_baixados()
arquivos_base = coleta_nomes_imagens()
tabela = compara_imagens(arquivos_baixados)
#print(orb_similarity)
